from fastapi import FastAPI
from fastapi.responses import FileResponse
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

logger = logging.getLogger(__name__)

app = FastAPI()


def get_db():
    return mysql.connector.connect(
        host="127.0.0.1",
        user="root",
        password="@Shayparmo24518",
        database="xlab_webdashboard"
    )

@app.get("/")
def home():
    return FileResponse(
        "templates/index.html",
        media_type="text/html"
    )

@app.get("/tickets")
def get_tickets():
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tickets")
    tickets = cursor.fetchall()

    logger.info("Tickets opgehaald: %s", len(tickets))

    cursor.close()
    db.close()

    return tickets
@app.get("/tickets/statussen")
def get_statussen():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT DISTINCT status
        FROM tickets
        ORDER BY status
    """)

    statussen = cursor.fetchall()

    cursor.close()
    db.close()

    return [status[0] for status in statussen]
@app.get("/tickets/status-count")
def status_count():
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            status,
            COUNT(*) AS aantal
        FROM tickets
        GROUP BY status
        ORDER BY status
    """)

    resultaten = cursor.fetchall()

    cursor.close()
    db.close()

    return resultaten
@app.get("/tickets/search")
def search_tickets(
    status: str | None = None,
    organisatie: str | None = None,
    zoekterm: str | None = None
):
    db = get_db()
    cursor = db.cursor(dictionary=True)

    query = "SELECT * FROM tickets WHERE 1=1"
    waarden = []

    if status:
        query += " AND TRIM(status) = TRIM(%s)"
        waarden.append(status)

    if organisatie:
        query += " AND TRIM(organisatie) = TRIM(%s)"
        waarden.append(organisatie)

    if zoekterm:
        query += """
            AND (
                ticketnr LIKE %s
                OR onderwerp LIKE %s
                OR toelichting LIKE %s
            )
        """

        zoek = f"%{zoekterm}%"
        waarden.extend([zoek, zoek, zoek])

    query += " ORDER BY id DESC"

    cursor.execute(query, tuple(waarden))
    tickets = cursor.fetchall()

    logger.info(
        "Ticketzoekopdracht uitgevoerd - status=%s, organisatie=%s, zoekterm=%s, resultaten=%s",
        status,
        organisatie,
        zoekterm,
        len(tickets)
    )

    cursor.close()
    db.close()

    return tickets

@app.get("/tickets-page")
def tickets_page():
    return FileResponse(
        "templates/tickets.html",
        media_type="text/html"
    )




from pydantic import BaseModel
from datetime import date

class TicketCreate(BaseModel):
    dossier_id: str
    status: str
    ticketnr:str
    soort_ticket: str
    aangemeld_door: str
    onderwerp: str
    omgeving: str
    prioriteit: str
    impact: str
    datum: date
    toelichting: str
    organisatie: str
    created_by: str
    modified_by: str

@app.get("/tickets/export")
def export_tickets():
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tickets ORDER BY id DESC")
    tickets = cursor.fetchall()

    cursor.close()
    db.close()

    if not tickets:
        return {"message": "Geen tickets gevonden"}

    # Nieuw Excel-bestand maken
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Kolommen uit de database
    kolommen = list(tickets[0].keys())

    # Header schrijven
    for kolom_nummer, kolom in enumerate(kolommen, start=1):
        cel = ws.cell(
            row=1,
            column=kolom_nummer,
            value=kolom
        )

        cel.font = Font(bold=True)
        cel.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )
        cel.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Tickets schrijven
    for rij_nummer, ticket in enumerate(tickets, start=2):
        for kolom_nummer, kolom in enumerate(kolommen, start=1):
            waarde = ticket.get(kolom)

            if waarde is not None:
                waarde = str(waarde)

            cel = ws.cell(
                row=rij_nummer,
                column=kolom_nummer,
                value=waarde
            )

            cel.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # Kolombreedtes automatisch instellen
    for kolom_nummer, kolom in enumerate(kolommen, start=1):
        maximale_lengte = len(str(kolom))

        for rij in range(2, ws.max_row + 1):
            waarde = ws.cell(
                row=rij,
                column=kolom_nummer
            ).value

            if waarde is not None:
                maximale_lengte = max(
                    maximale_lengte,
                    len(str(waarde))
                )

        # Niet extreem breed maken
        breedte = min(max(maximale_lengte + 2, 12), 45)

        ws.column_dimensions[
            get_column_letter(kolom_nummer)
        ].width = breedte

    # Eerste rij vastzetten
    ws.freeze_panes = "A2"

    # Filter op alle kolommen
    ws.auto_filter.ref = ws.dimensions

    # Excel-bestand opslaan
    bestand = os.path.abspath("tickets_export.xlsx")

    wb.save(bestand)

    return FileResponse(
        path=bestand,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="tickets_export.xlsx")
@app.post("/tickets")
def create_ticket(ticket: TicketCreate):
    db = get_db()
    cursor = db.cursor()

    cursor.execute(
        """
        INSERT INTO tickets (
            dossier_id,
            status,
            ticketnr,
            soort_ticket,
            aangemeld_door,
            onderwerp,
            omgeving,
            prio,
            impact,
            datum,
            toelichting,
            organisatie,
            created_by,
            modified_by
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            ticket.dossier_id,
            ticket.status,
            ticket.ticketnr,
            ticket.soort_ticket,
            ticket.aangemeld_door,
            ticket.onderwerp,
            ticket.omgeving,
            ticket.prioriteit,
            ticket.impact,
            ticket.datum,
            ticket.toelichting,
            ticket.organisatie,
            ticket.created_by,
            ticket.modified_by
        )
    )

    db.commit()

    new_id = cursor.lastrowid

    cursor.close()
    db.close()

    return {
        "message": "Ticket succesvol aangemaakt",
        "id": new_id
    }
class TicketUpdate(BaseModel):
    dossier_id: str
    status: str
    ticketnr: str
    soort_ticket: str
    aangemeld_door: str
    onderwerp: str
    omgeving: str
    prioriteit: str
    impact: str
    datum: date
    toelichting: str
    organisatie: str
    created_by: str
    modified_by: str

@app.put("/tickets/{ticket_id}")
def update_ticket(ticket_id: int, ticket: TicketUpdate):
    db = get_db()
    cursor = db.cursor()

    cursor.execute(
        """
        UPDATE tickets
        SET
            dossier_id = %s,
            status = %s,
            ticketnr = %s,
            soort_ticket = %s,
            aangemeld_door = %s,
            onderwerp = %s,
            omgeving = %s,
            prio = %s,
            impact = %s,
            datum = %s,
            toelichting = %s,
            organisatie = %s,
            created_by = %s,
            modified_by = %s
        WHERE id = %s
        """,
        (
            ticket.dossier_id,
            ticket.status,
            ticket.ticketnr,
            ticket.soort_ticket,
            ticket.aangemeld_door,
            ticket.onderwerp,
            ticket.omgeving,
            ticket.prioriteit,
            ticket.impact,
            ticket.datum,
            ticket.toelichting,
            ticket.organisatie,
            ticket.created_by,
            ticket.modified_by,
            ticket_id
        )
    )

    db.commit()

    if cursor.rowcount == 0:
        cursor.close()
        db.close()
        return {"message": "Ticket niet gevonden"}

    cursor.close()
    db.close()

    return {
        "message": "Ticket succesvol gewijzigd",
        "id": ticket_id
    }
@app.delete("/tickets/{ticket_id}")
def delete_ticket(ticket_id: int):
    db = get_db()
    cursor = db.cursor()

    cursor.execute(
        "DELETE FROM tickets WHERE id = %s",
        (ticket_id,)
    )

    db.commit()

    if cursor.rowcount == 0:
        cursor.close()
        db.close()
        return {"message": "Ticket niet gevonden"}

    cursor.close()
    db.close()

    return {
        "message": "Ticket succesvol verwijderd",
        "id": ticket_id
    }
